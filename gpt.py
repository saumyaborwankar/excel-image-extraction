from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import io
import os
import zipfile
import xml.etree.ElementTree as ET


def emu_to_pixels(emu):
    return emu / 9525 if emu else 0


def cell_to_pixels(ws, row, col):
    """Convert cell position to absolute pixel coordinates."""
    x = 0
    for c in range(1, col):
        col_width = ws.column_dimensions[get_column_letter(c)].width
        if col_width is None:
            col_width = 8.43
        x += col_width * 7

    y = 0
    for r in range(1, row):
        row_height = ws.row_dimensions[r].height
        if row_height is None:
            row_height = 15
        y += row_height * 1.33

    return x, y


def calculate_bottom_right_cell(
    ws,
    start_row,
    start_col,
    start_offset_x,
    start_offset_y,
    img_width_px,
    img_height_px,
):
    """
    Calculate the bottom-right cell based on image dimensions.
    Returns (cell_coordinate, offset_x, offset_y)
    """
    current_row = start_row
    current_col = start_col
    remaining_width = start_offset_x + img_width_px
    remaining_height = start_offset_y + img_height_px

    # Traverse columns
    while remaining_width > 0:
        col_width_px = ws.column_dimensions[get_column_letter(current_col)].width
        if col_width_px is None:
            col_width_px = 8.43
        col_width_px *= 7

        if remaining_width <= col_width_px:
            break
        remaining_width -= col_width_px
        current_col += 1

    # Traverse rows
    while remaining_height > 0:
        row_height_px = ws.row_dimensions[current_row].height
        if row_height_px is None:
            row_height_px = 15
        row_height_px *= 1.33

        if remaining_height <= row_height_px:
            break
        remaining_height -= row_height_px
        current_row += 1

    cell_coord = ws.cell(row=current_row, column=current_col).coordinate
    return cell_coord, remaining_width, remaining_height


def rectangles_overlap(rect1, rect2):
    """Check if two rectangles overlap. Rectangles are (x, y, width, height)."""
    x1, y1, w1, h1 = rect1
    x2, y2, w2, h2 = rect2

    return not (
        x1 + w1 <= x2  # rect1 is left of rect2
        or x2 + w2 <= x1  # rect2 is left of rect1
        or y1 + h1 <= y2  # rect1 is above rect2
        or y2 + h2 <= y1
    )  # rect2 is above rect1


def is_overlay(base_rect, overlay_rect):
    """
    Check if overlay_rect is completely within base_rect.
    This indicates overlay_rect is an overlay on top of base_rect.
    """
    bx, by, bw, bh = base_rect
    ox, oy, ow, oh = overlay_rect

    return ox >= bx and oy >= by and ox + ow <= bx + bw and oy + oh <= by + bh


def extract_shapes_from_excel(file_path, sheet_name):
    """
    Extract shape information from Excel file by parsing the drawing XML.
    Returns a list of shape dictionaries with position, size, and properties.
    """
    shapes = []

    try:
        with zipfile.ZipFile(file_path, "r") as zip_ref:
            # Find drawing files
            drawing_files = [
                f
                for f in zip_ref.namelist()
                if "xl/drawings/drawing" in f and f.endswith(".xml") and "rels" not in f
            ]

            for drawing_file in drawing_files:
                content = zip_ref.read(drawing_file)
                root = ET.fromstring(content)

                # Define namespaces
                ns = {
                    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                }

                # Process two-cell anchors
                for anchor in root.findall(".//xdr:twoCellAnchor", ns):
                    sp = anchor.find(".//xdr:sp", ns)
                    if sp is not None:  # It's a shape, not an image
                        shape_info = {"type": "shape", "anchor_type": "two_cell"}

                        # Get shape name
                        nv_sp_pr = sp.find(".//xdr:nvSpPr", ns)
                        if nv_sp_pr is not None:
                            c_nv_pr = nv_sp_pr.find(".//xdr:cNvPr", ns)
                            if c_nv_pr is not None:
                                shape_info["name"] = c_nv_pr.get("name", "Unknown")

                        # Get shape type (geometry) and colors
                        sp_pr = sp.find(".//xdr:spPr", ns)
                        if sp_pr is not None:
                            # Check for preset geometry
                            prst_geom = sp_pr.find(".//a:prstGeom", ns)
                            if prst_geom is not None:
                                shape_info["geometry"] = prst_geom.get("prst", "rect")
                            else:
                                shape_info["geometry"] = "rect"

                            # Check for no fill first (takes precedence)
                            no_fill = sp_pr.find(".//a:noFill", ns)
                            if no_fill is not None:
                                shape_info["fill_color"] = None
                            else:
                                # Get fill color
                                solid_fill = sp_pr.find(".//a:solidFill", ns)
                                if solid_fill is not None:
                                    srgb_clr = solid_fill.find(".//a:srgbClr", ns)
                                    if srgb_clr is not None:
                                        fill_color = srgb_clr.get("val", "FFFF00")
                                        shape_info["fill_color"] = f"#{fill_color}"

                                        # Get alpha/transparency
                                        alpha = srgb_clr.find("a:alpha", ns)
                                        if alpha is not None:
                                            alpha_val = int(alpha.get("val", "100000"))
                                            shape_info["fill_alpha"] = int(
                                                255 * alpha_val / 100000
                                            )
                                        else:
                                            shape_info["fill_alpha"] = 255
                                    else:
                                        shape_info["fill_color"] = None
                                else:
                                    shape_info["fill_color"] = None

                            # Get outline color
                            ln = sp_pr.find(".//a:ln", ns)
                            if ln is not None:
                                width = ln.get("w", "0")
                                shape_info["outline_width"] = (
                                    int(width) / 12700 if width != "0" else 0
                                )

                                solid_fill = ln.find(".//a:solidFill", ns)
                                if solid_fill is not None:
                                    srgb_clr = solid_fill.find(".//a:srgbClr", ns)
                                    if srgb_clr is not None:
                                        outline_color = srgb_clr.get("val", "000000")
                                        shape_info["outline_color"] = (
                                            f"#{outline_color}"
                                        )
                                    else:
                                        shape_info["outline_color"] = None
                                else:
                                    shape_info["outline_color"] = None
                            else:
                                shape_info["outline_width"] = 1
                                shape_info["outline_color"] = None
                        else:
                            shape_info["geometry"] = "rect"
                            shape_info["fill_color"] = None
                            shape_info["outline_color"] = None

                        # Get text content and font properties
                        text_body = sp.find(".//xdr:txBody", ns)
                        if text_body is not None:
                            text_content = []
                            for t in text_body.findall(".//a:t", ns):
                                if t.text:
                                    text_content.append(t.text)
                            shape_info["text"] = " ".join(text_content)

                            # Get font size from text run properties
                            font_size = None
                            for para in text_body.findall(".//a:p", ns):
                                # Check run properties first
                                for run in para.findall(".//a:r", ns):
                                    rpr = run.find("a:rPr", ns)
                                    if rpr is not None:
                                        sz = rpr.get("sz")
                                        if sz:
                                            font_size = (
                                                int(sz) / 100
                                            )  # Convert from hundredths of points
                                            break

                                # If no run properties, check default run properties
                                if font_size is None:
                                    def_rpr = para.find(".//a:defRPr", ns)
                                    if def_rpr is not None:
                                        sz = def_rpr.get("sz")
                                        if sz:
                                            font_size = int(sz) / 100

                                if font_size:
                                    break

                            # Store font size if found, otherwise use default
                            shape_info["font_size_pt"] = (
                                font_size if font_size else 11.0
                            )

                        # Get position
                        from_cell = anchor.find(".//xdr:from", ns)
                        to_cell = anchor.find(".//xdr:to", ns)

                        if from_cell is not None:
                            col = from_cell.find("xdr:col", ns)
                            row = from_cell.find("xdr:row", ns)
                            col_off = from_cell.find("xdr:colOff", ns)
                            row_off = from_cell.find("xdr:rowOff", ns)

                            if col is not None and row is not None:
                                shape_info["from_col"] = int(col.text)
                                shape_info["from_row"] = int(row.text)
                                shape_info["from_col_off"] = (
                                    int(col_off.text) if col_off is not None else 0
                                )
                                shape_info["from_row_off"] = (
                                    int(row_off.text) if row_off is not None else 0
                                )

                        if to_cell is not None:
                            col = to_cell.find("xdr:col", ns)
                            row = to_cell.find("xdr:row", ns)
                            col_off = to_cell.find("xdr:colOff", ns)
                            row_off = to_cell.find("xdr:rowOff", ns)

                            if col is not None and row is not None:
                                shape_info["to_col"] = int(col.text)
                                shape_info["to_row"] = int(row.text)
                                shape_info["to_col_off"] = (
                                    int(col_off.text) if col_off is not None else 0
                                )
                                shape_info["to_row_off"] = (
                                    int(row_off.text) if row_off is not None else 0
                                )

                        shapes.append(shape_info)

                # Process one-cell anchors
                for anchor in root.findall(".//xdr:oneCellAnchor", ns):
                    sp = anchor.find(".//xdr:sp", ns)
                    if sp is not None:  # It's a shape, not an image
                        shape_info = {"type": "shape", "anchor_type": "one_cell"}

                        # Get shape name
                        nv_sp_pr = sp.find(".//xdr:nvSpPr", ns)
                        if nv_sp_pr is not None:
                            c_nv_pr = nv_sp_pr.find(".//xdr:cNvPr", ns)
                            if c_nv_pr is not None:
                                shape_info["name"] = c_nv_pr.get("name", "Unknown")

                        # Get shape type (geometry) and colors
                        sp_pr = sp.find(".//xdr:spPr", ns)
                        if sp_pr is not None:
                            # Check for preset geometry
                            prst_geom = sp_pr.find(".//a:prstGeom", ns)
                            if prst_geom is not None:
                                shape_info["geometry"] = prst_geom.get("prst", "rect")
                            else:
                                shape_info["geometry"] = "rect"

                            # Check for no fill first (takes precedence)
                            no_fill = sp_pr.find(".//a:noFill", ns)
                            if no_fill is not None:
                                shape_info["fill_color"] = None
                            else:
                                # Get fill color
                                solid_fill = sp_pr.find(".//a:solidFill", ns)
                                if solid_fill is not None:
                                    srgb_clr = solid_fill.find(".//a:srgbClr", ns)
                                    if srgb_clr is not None:
                                        fill_color = srgb_clr.get("val", "FFFF00")
                                        shape_info["fill_color"] = f"#{fill_color}"

                                        # Get alpha/transparency
                                        alpha = srgb_clr.find("a:alpha", ns)
                                        if alpha is not None:
                                            alpha_val = int(alpha.get("val", "100000"))
                                            shape_info["fill_alpha"] = int(
                                                255 * alpha_val / 100000
                                            )
                                        else:
                                            shape_info["fill_alpha"] = 255
                                    else:
                                        shape_info["fill_color"] = None
                                else:
                                    shape_info["fill_color"] = None

                            # Get outline color
                            ln = sp_pr.find(".//a:ln", ns)
                            if ln is not None:
                                width = ln.get("w", "0")
                                shape_info["outline_width"] = (
                                    int(width) / 12700 if width != "0" else 0
                                )

                                solid_fill = ln.find(".//a:solidFill", ns)
                                if solid_fill is not None:
                                    srgb_clr = solid_fill.find(".//a:srgbClr", ns)
                                    if srgb_clr is not None:
                                        outline_color = srgb_clr.get("val", "000000")
                                        shape_info["outline_color"] = (
                                            f"#{outline_color}"
                                        )
                                    else:
                                        shape_info["outline_color"] = None
                                else:
                                    shape_info["outline_color"] = None
                            else:
                                shape_info["outline_width"] = 1
                                shape_info["outline_color"] = None
                        else:
                            shape_info["geometry"] = "rect"
                            shape_info["fill_color"] = None
                            shape_info["outline_color"] = None

                        # Get text content and font properties
                        text_body = sp.find(".//xdr:txBody", ns)
                        if text_body is not None:
                            text_content = []
                            for t in text_body.findall(".//a:t", ns):
                                if t.text:
                                    text_content.append(t.text)
                            shape_info["text"] = " ".join(text_content)

                            # Get font size from text run properties
                            font_size = None
                            for para in text_body.findall(".//a:p", ns):
                                # Check run properties first
                                for run in para.findall(".//a:r", ns):
                                    rpr = run.find("a:rPr", ns)
                                    if rpr is not None:
                                        sz = rpr.get("sz")
                                        if sz:
                                            font_size = (
                                                int(sz) / 100
                                            )  # Convert from hundredths of points
                                            break

                                # If no run properties, check default run properties
                                if font_size is None:
                                    def_rpr = para.find(".//a:defRPr", ns)
                                    if def_rpr is not None:
                                        sz = def_rpr.get("sz")
                                        if sz:
                                            font_size = int(sz) / 100

                                if font_size:
                                    break

                            # Store font size if found, otherwise use default
                            shape_info["font_size_pt"] = (
                                font_size if font_size else 11.0
                            )

                        # Get position
                        from_cell = anchor.find(".//xdr:from", ns)
                        if from_cell is not None:
                            col = from_cell.find("xdr:col", ns)
                            row = from_cell.find("xdr:row", ns)
                            col_off = from_cell.find("xdr:colOff", ns)
                            row_off = from_cell.find("xdr:rowOff", ns)

                            if col is not None and row is not None:
                                shape_info["from_col"] = int(col.text)
                                shape_info["from_row"] = int(row.text)
                                shape_info["from_col_off"] = (
                                    int(col_off.text) if col_off is not None else 0
                                )
                                shape_info["from_row_off"] = (
                                    int(row_off.text) if row_off is not None else 0
                                )

                        # Get extent (size)
                        ext = anchor.find(".//xdr:ext", ns)
                        if ext is not None:
                            shape_info["width_emu"] = int(ext.get("cx", "0"))
                            shape_info["height_emu"] = int(ext.get("cy", "0"))

                        shapes.append(shape_info)

    except Exception as e:
        print(f"[Warning] Could not extract shapes: {e}")

    return shapes


file_path = "sample.xlsx"
output_dir = "images_with_positions"
overlay_dir = "images_with_overlays"
os.makedirs(output_dir, exist_ok=True)
os.makedirs(overlay_dir, exist_ok=True)

wb = load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Extract shapes from the Excel file
    print(f"\n{'='*60}")
    print(f"Processing sheet: {sheet_name}")
    print("=" * 60)

    shapes = extract_shapes_from_excel(file_path, sheet_name)
    print(f"\n🔷 Found {len(shapes)} shape(s) in drawing XML")

    # Process shapes to calculate their absolute positions
    shape_data = []
    for idx, shape in enumerate(shapes, 1):
        abs_x = abs_y = 0
        width_px = height_px = 0

        if "from_col" in shape and "from_row" in shape:
            cell_x, cell_y = cell_to_pixels(
                ws, shape["from_row"] + 1, shape["from_col"] + 1
            )
            abs_x = cell_x + emu_to_pixels(shape.get("from_col_off", 0))
            abs_y = cell_y + emu_to_pixels(shape.get("from_row_off", 0))

            if (
                shape["anchor_type"] == "two_cell"
                and "to_col" in shape
                and "to_row" in shape
            ):
                to_cell_x, to_cell_y = cell_to_pixels(
                    ws, shape["to_row"] + 1, shape["to_col"] + 1
                )
                to_x = to_cell_x + emu_to_pixels(shape.get("to_col_off", 0))
                to_y = to_cell_y + emu_to_pixels(shape.get("to_row_off", 0))
                width_px = to_x - abs_x
                height_px = to_y - abs_y
            elif "width_emu" in shape and "height_emu" in shape:
                width_px = emu_to_pixels(shape["width_emu"])
                height_px = emu_to_pixels(shape["height_emu"])

        shape["abs_x"] = abs_x
        shape["abs_y"] = abs_y
        shape["width"] = width_px
        shape["height"] = height_px
        shape_data.append(shape)

        print(f"  Shape {idx}: {shape.get('name', 'Unknown')}")
        print(f"    Geometry: {shape.get('geometry', 'rect')}")
        if "text" in shape:
            print(f"    Text: {shape['text']}")
        print(f"    Position: ({abs_x:.1f}, {abs_y:.1f})")
        print(f"    Size: {width_px:.1f}x{height_px:.1f} px")

    # First pass: collect all image data
    image_data = []

    for idx, image in enumerate(ws._images, start=1):
        anchor = image.anchor
        top_left_cell = "Unknown"
        bottom_right_cell = "Unknown"
        offset_x = offset_y = 0
        end_offset_x = end_offset_y = 0
        abs_x = abs_y = 0
        img_width_px = img_height_px = 0

        try:
            if hasattr(anchor, "_from") and hasattr(anchor, "to"):
                start = anchor._from
                top_left_cell = ws.cell(
                    row=start.row + 1, column=start.col + 1
                ).coordinate
                offset_x = emu_to_pixels(start.colOff)
                offset_y = emu_to_pixels(start.rowOff)

                # Calculate absolute position
                cell_x, cell_y = cell_to_pixels(ws, start.row + 1, start.col + 1)
                abs_x = cell_x + offset_x
                abs_y = cell_y + offset_y

                end = anchor.to
                bottom_right_cell = ws.cell(
                    row=end.row + 1, column=end.col + 1
                ).coordinate
                end_offset_x = emu_to_pixels(end.colOff)
                end_offset_y = emu_to_pixels(end.rowOff)

                # Calculate dimensions
                end_cell_x, end_cell_y = cell_to_pixels(ws, end.row + 1, end.col + 1)
                img_width_px = (end_cell_x + end_offset_x) - abs_x
                img_height_px = (end_cell_y + end_offset_y) - abs_y

            elif hasattr(anchor, "_from"):
                start = anchor._from
                top_left_cell = ws.cell(
                    row=start.row + 1, column=start.col + 1
                ).coordinate
                offset_x = emu_to_pixels(start.colOff)
                offset_y = emu_to_pixels(start.rowOff)

                # Calculate absolute position
                cell_x, cell_y = cell_to_pixels(ws, start.row + 1, start.col + 1)
                abs_x = cell_x + offset_x
                abs_y = cell_y + offset_y

                img_width_px = (
                    emu_to_pixels(anchor.ext.cx) if hasattr(anchor, "ext") else 0
                )
                img_height_px = (
                    emu_to_pixels(anchor.ext.cy) if hasattr(anchor, "ext") else 0
                )

                if img_width_px > 0 and img_height_px > 0:
                    bottom_right_cell, end_offset_x, end_offset_y = (
                        calculate_bottom_right_cell(
                            ws,
                            start.row + 1,
                            start.col + 1,
                            offset_x,
                            offset_y,
                            img_width_px,
                            img_height_px,
                        )
                    )
                else:
                    bottom_right_cell = "Unknown"

            elif isinstance(anchor, str):
                top_left_cell = anchor

        except Exception as e:
            print(f"[Warning] Could not parse anchor for image {idx}: {e}")

        if top_left_cell != "Unknown" and bottom_right_cell != "Unknown":
            img_filename = f"{top_left_cell}-{bottom_right_cell}.png"
        else:
            img_filename = f"{sheet_name}_image_{idx}.png"

        # Save individual image
        img_bytes = image._data()
        img_path = os.path.join(output_dir, img_filename)
        with open(img_path, "wb") as f:
            f.write(img_bytes)

        # Load image for overlay detection
        try:
            pil_image = Image.open(io.BytesIO(img_bytes))
        except Exception as e:
            print(f"[Warning] Could not load image {img_filename}: {e}")
            pil_image = None

        image_data.append(
            {
                "idx": idx,
                "filename": img_filename,
                "top_left_cell": top_left_cell,
                "bottom_right_cell": bottom_right_cell,
                "offset_x": offset_x,
                "offset_y": offset_y,
                "end_offset_x": end_offset_x,
                "end_offset_y": end_offset_y,
                "abs_x": abs_x,
                "abs_y": abs_y,
                "width": img_width_px,
                "height": img_height_px,
                "pil_image": pil_image,
                "bytes": img_bytes,
            }
        )

        print(f"Sheet: {sheet_name}")
        print(f"  ➜ Image: {img_filename}")
        print(f"  ➜ Top-Left Cell: {top_left_cell}")
        print(f"  ➜ Top-Left Offset (pixels): x={offset_x:.1f}, y={offset_y:.1f}")
        print(f"  ➜ Bottom-Right Cell: {bottom_right_cell}")
        print(
            f"  ➜ Bottom-Right Offset (pixels): x={end_offset_x:.1f}, y={end_offset_y:.1f}"
        )
        print(f"  ➜ Dimensions: {img_width_px:.1f}x{img_height_px:.1f} px")
        print("-" * 50)

    # Second pass: detect overlays (images and shapes) and create composites
    print(f"\n🔍 Checking for overlays in sheet '{sheet_name}'...")
    print(f"\nDebug: Found {len(image_data)} image(s) and {len(shape_data)} shape(s)")

    # Print all image positions for debugging
    for img in image_data:
        print(
            f"  - Image: {img['filename']}: pos=({img['abs_x']:.1f}, {img['abs_y']:.1f}), size=({img['width']:.1f}x{img['height']:.1f})"
        )

    for i, base_img in enumerate(image_data):
        if (
            base_img["pil_image"] is None
            or base_img["width"] <= 0
            or base_img["height"] <= 0
        ):
            continue

        base_rect = (
            base_img["abs_x"],
            base_img["abs_y"],
            base_img["width"],
            base_img["height"],
        )

        overlays = []
        shape_overlays = []

        # Check for overlapping images
        for j, overlay_img in enumerate(image_data):
            if i == j or overlay_img["pil_image"] is None:
                continue

            overlay_rect = (
                overlay_img["abs_x"],
                overlay_img["abs_y"],
                overlay_img["width"],
                overlay_img["height"],
            )

            # Check if they overlap at all first
            if rectangles_overlap(base_rect, overlay_rect):
                print(
                    f"\n🔄 Images overlap: {base_img['filename']} and {overlay_img['filename']}"
                )
                print(
                    f"   Base: x={base_rect[0]:.1f}, y={base_rect[1]:.1f}, w={base_rect[2]:.1f}, h={base_rect[3]:.1f}"
                )
                print(
                    f"   Overlay: x={overlay_rect[0]:.1f}, y={overlay_rect[1]:.1f}, w={overlay_rect[2]:.1f}, h={overlay_rect[3]:.1f}"
                )

                # Check if overlay_img is on top of base_img (completely within)
                if is_overlay(base_rect, overlay_rect):
                    print(f"   ✓ Is an overlay (completely within base)")
                    overlays.append(("image", j, overlay_img))
                # Check if base is within overlay (reverse case)
                elif is_overlay(overlay_rect, base_rect):
                    print(
                        f"   ⚠ Base is within overlay (skipping - will be handled when overlay is base)"
                    )
                else:
                    # If they overlap but neither is completely within the other,
                    # treat smaller image as overlay on larger image
                    base_area = base_rect[2] * base_rect[3]
                    overlay_area = overlay_rect[2] * overlay_rect[3]
                    if overlay_area < base_area:
                        print(f"   ✓ Partial overlap - treating smaller as overlay")
                        overlays.append(("image", j, overlay_img))
                    else:
                        print(
                            f"   ⚠ Partial overlap - but overlay is larger (skipping)"
                        )

        # Check for shapes overlaying this image
        for shape in shape_data:
            if shape["width"] <= 0 or shape["height"] <= 0:
                continue

            shape_rect = (
                shape["abs_x"],
                shape["abs_y"],
                shape["width"],
                shape["height"],
            )

            if rectangles_overlap(base_rect, shape_rect):
                print(
                    f"\n🔷 Shape overlays image: {shape.get('name', 'Unknown')} on {base_img['filename']}"
                )
                print(
                    f"   Base: x={base_rect[0]:.1f}, y={base_rect[1]:.1f}, w={base_rect[2]:.1f}, h={base_rect[3]:.1f}"
                )
                print(
                    f"   Shape: x={shape_rect[0]:.1f}, y={shape_rect[1]:.1f}, w={shape_rect[2]:.1f}, h={shape_rect[3]:.1f}"
                )
                shape_overlays.append(shape)

        if overlays or shape_overlays:
            print(
                f"\n📌 Found {len(overlays)} image overlay(s) and {len(shape_overlays)} shape overlay(s) on {base_img['filename']}"
            )

            # Create composite image
            composite = base_img["pil_image"].convert("RGBA")
            draw = ImageDraw.Draw(composite)

            # Debug: show image dimensions
            actual_img_width = base_img["pil_image"].width
            actual_img_height = base_img["pil_image"].height
            print(
                f"   Image dimensions: calculated={base_img['width']:.1f}x{base_img['height']:.1f}, actual={actual_img_width}x{actual_img_height}"
            )
            scale_x = (
                actual_img_width / base_img["width"] if base_img["width"] > 0 else 1
            )
            scale_y = (
                actual_img_height / base_img["height"] if base_img["height"] > 0 else 1
            )
            print(f"   Scaling factors: x={scale_x:.3f}, y={scale_y:.3f}")

            # Add image overlays
            for overlay_type, overlay_idx, overlay_img in overlays:
                # Calculate relative position
                rel_x = int(overlay_img["abs_x"] - base_img["abs_x"])
                rel_y = int(overlay_img["abs_y"] - base_img["abs_y"])

                overlay_pil = overlay_img["pil_image"].convert("RGBA")
                composite.paste(overlay_pil, (rel_x, rel_y), overlay_pil)

                print(
                    f"   ↳ Image overlay: {overlay_img['filename']} at ({rel_x}, {rel_y})"
                )

            # Add shape overlays (render according to geometry type)
            for shape in shape_overlays:
                # Calculate position in the coordinate system
                rel_x = shape["abs_x"] - base_img["abs_x"]
                rel_y = shape["abs_y"] - base_img["abs_y"]

                # Scale to match actual image dimensions
                # The base_img['width'] and ['height'] are in our calculated pixels
                # but the actual image might have different dimensions
                actual_img_width = base_img["pil_image"].width
                actual_img_height = base_img["pil_image"].height

                scale_x = (
                    actual_img_width / base_img["width"] if base_img["width"] > 0 else 1
                )
                scale_y = (
                    actual_img_height / base_img["height"]
                    if base_img["height"] > 0
                    else 1
                )

                # Apply scaling
                rel_x = int(rel_x * scale_x)
                rel_y = int(rel_y * scale_y)
                rel_x2 = rel_x + int(shape["width"] * scale_x)
                rel_y2 = rel_y + int(shape["height"] * scale_y)

                # Create a layer for this shape
                shape_layer = Image.new("RGBA", composite.size, (0, 0, 0, 0))
                shape_draw = ImageDraw.Draw(shape_layer)

                geometry = shape.get("geometry", "rect")

                # Get colors from shape data
                fill_color = None
                if shape.get("fill_color"):
                    hex_color = shape["fill_color"].lstrip("#")
                    r, g, b = tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))
                    alpha = shape.get("fill_alpha", 255)
                    fill_color = (r, g, b, alpha)

                outline_color = None
                outline_width = 1
                if shape.get("outline_color"):
                    hex_color = shape["outline_color"].lstrip("#")
                    r, g, b = tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))
                    outline_color = (r, g, b, 255)
                    outline_width = max(1, int(shape.get("outline_width", 1) * scale_x))

                # Draw shape based on geometry type
                if geometry == "ellipse":
                    # Draw ellipse/circle
                    shape_draw.ellipse(
                        [rel_x, rel_y, rel_x2, rel_y2],
                        fill=fill_color,
                        outline=outline_color,
                        width=outline_width,
                    )
                elif geometry in ["rect", "rectangle"]:
                    # Draw rectangle
                    shape_draw.rectangle(
                        [rel_x, rel_y, rel_x2, rel_y2],
                        fill=fill_color,
                        outline=outline_color,
                        width=outline_width,
                    )
                elif geometry == "roundRect":
                    # Draw rounded rectangle
                    shape_draw.rounded_rectangle(
                        [rel_x, rel_y, rel_x2, rel_y2],
                        radius=int(10 * scale_x),
                        fill=fill_color,
                        outline=outline_color,
                        width=outline_width,
                    )
                elif geometry in ["triangle", "rtTriangle"]:
                    # Draw triangle
                    shape_draw.polygon(
                        [
                            (
                                rel_x + int(shape["width"] * scale_x) / 2,
                                rel_y,
                            ),  # Top center
                            (rel_x, rel_y2),  # Bottom left
                            (rel_x2, rel_y2),
                        ],  # Bottom right
                        fill=fill_color,
                        outline=outline_color,
                    )
                else:
                    # Default to rectangle for unknown shapes
                    shape_draw.rectangle(
                        [rel_x, rel_y, rel_x2, rel_y2],
                        fill=fill_color,
                        outline=outline_color,
                        width=outline_width,
                    )

                # Add text if present
                if "text" in shape and shape["text"]:
                    try:
                        # Get font size and scale it
                        font_size_pt = shape.get("font_size_pt", 11.0)
                        # Scale font size by the average of scale_x and scale_y
                        scaled_font_size = int(font_size_pt * (scale_x + scale_y) / 2)
                        print(
                            f"   [Debug] Text '{shape['text']}': original font size={font_size_pt}pt, scaled={scaled_font_size}px"
                        )

                        # Try to use a TrueType font, fall back to default if not available
                        try:
                            # Try common system fonts
                            font = ImageFont.truetype("Arial.ttf", scaled_font_size)
                        except:
                            try:
                                font = ImageFont.truetype(
                                    "/System/Library/Fonts/Supplemental/Arial.ttf",
                                    scaled_font_size,
                                )
                            except:
                                try:
                                    font = ImageFont.truetype(
                                        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                                        scaled_font_size,
                                    )
                                except:
                                    # Fall back to default font if no TrueType font found
                                    font = ImageFont.load_default()

                        text_bbox = shape_draw.textbbox(
                            (rel_x, rel_y), shape["text"], font=font
                        )
                        text_width = text_bbox[2] - text_bbox[0]
                        text_height = text_bbox[3] - text_bbox[1]
                        shape_width_scaled = int(shape["width"] * scale_x)
                        shape_height_scaled = int(shape["height"] * scale_y)
                        text_x = rel_x + (shape_width_scaled - text_width) / 2
                        text_y = rel_y + (shape_height_scaled - text_height) / 2
                        shape_draw.text(
                            (text_x, text_y),
                            shape["text"],
                            fill=(0, 0, 0, 255),
                            font=font,
                        )
                    except Exception as e:
                        print(f"   [Warning] Could not render text: {e}")

                composite = Image.alpha_composite(composite, shape_layer)
                print(
                    f"   ↳ Shape overlay: {shape.get('name', 'Unknown')} ({geometry}) at ({rel_x}, {rel_y})"
                )

            # Save composite as JPG
            composite_filename = (
                f"{base_img['filename'].replace('.png', '')}_with_overlays.jpg"
            )
            composite_path = os.path.join(overlay_dir, composite_filename)

            # Convert to RGB for JPEG
            composite_rgb = composite.convert("RGB")
            composite_rgb.save(composite_path, "JPEG", quality=95)

            print(f"   ✓ Saved composite: {composite_filename}")

print(f"\n✅ Individual images saved to: {output_dir}")
print(f"✅ Images with overlays saved to: {overlay_dir}")
